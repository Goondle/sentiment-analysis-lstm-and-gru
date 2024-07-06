import pickle
import string
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import os
import pandas as pd

import nltk
import pymorphy2
from nltk import SnowballStemmer
from nltk.corpus import stopwords
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing.sequence import pad_sequences
from keras import backend as K

# selected_model_path = ""
selected_model_path = None  #"C:/Users/user/PycharmProjects/sentimenLSTM/LSTM/LSTM_LEM/123/New folder/LSTM_LEM_model.keras"#None
selected_toki_path = None #"C:/Users/user/PycharmProjects/sentimenLSTM/LSTM/LSTM_LEM/123/New folder/lstm_lem_tokenizer.pickle"#None
selected_doc = None
selected_model_paths_for_model = []
selected_toki_path_for_model = []

lol = 0
if lol==1:
    selected_model_paths_for_model.append(
        "C:/Users/user/PycharmProjects/sentimenLSTM/models/GRU_LEM/GRU_LEM_model.keras")
    selected_model_paths_for_model.append(
        "C:/Users/user/PycharmProjects/sentimenLSTM/models/GRU_STEM/GRU_STEM_model.keras")
    selected_model_paths_for_model.append(
        "C:/Users/user/PycharmProjects/sentimenLSTM/models/LSTM_LEM/LSTM_LEM_model.keras")
    selected_model_paths_for_model.append(
        "C:/Users/user/PycharmProjects/sentimenLSTM/models/LSTM_STEM/LSTM_STEM_model.keras")
    selected_toki_path_for_model.append(
        "C:/Users/user/PycharmProjects/sentimenLSTM/models/GRU_LEM/gru_lem_tokenizer.pickle")
    selected_toki_path_for_model.append(
        "C:/Users/user/PycharmProjects/sentimenLSTM/models/GRU_STEM/gru_stem_tokenizer.pickle")
    selected_toki_path_for_model.append(
        "C:/Users/user/PycharmProjects/sentimenLSTM/models/LSTM_LEM/lstm_lem_tokenizer.pickle")
    selected_toki_path_for_model.append(
        "C:/Users/user/PycharmProjects/sentimenLSTM/models/LSTM_STEM/lstm_stem_tokenizer.pickle")
else:
    print("pupa")


def recall_m(y_true, y_pred):
    true_positives = K.sum(K.round(K.clip(y_true * y_pred, 0, 1)))
    possible_positives = K.sum(K.round(K.clip(y_true, 0, 1)))
    recall = true_positives / (possible_positives + K.epsilon())
    return recall
def precision_m(y_true, y_pred):
    true_positives = K.sum(K.round(K.clip(y_true * y_pred, 0, 1)))
    predicted_positives = K.sum(K.round(K.clip(y_pred, 0, 1)))
    precision = true_positives / (predicted_positives + K.epsilon())
    return precision
def f1_m(y_true, y_pred):
    precision = precision_m(y_true, y_pred)
    recall = recall_m(y_true, y_pred)
    return 2 * ((precision * recall) / (precision + recall + K.epsilon()))
morph = pymorphy2.MorphAnalyzer()
stop_words = set(stopwords.words("russian"))
# nltk.download('snowball_data')
stemmer = SnowballStemmer("russian")
def remove_digits_and_punctuation(text):
    text = text.lower()  # нижний регистр
    punctuation = string.punctuation + '«»'  # Создаем строку, содержащую все знаки препинания
    result = ''.join(' ' if char in punctuation or char.isdigit() else char for char in text)  # Удаляем все знаки препинания и цифры из текста, заменяя их на пробелы
    result = ' '.join(result.split())  # Удаляем лишние пробелы с помощью split и join
    result = result.replace('ё', 'е')  # заменяем букву "ё" на букву "е"     result = result.replace('€', 'деньги')
    return ' '.join(result.split())

def lemmatize(text):
    words = text.split()
    lemmatized_words = [morph.parse(word)[0].normal_form for word in words]
    lemmatized_text = ' '.join(lemmatized_words)
    return lemmatized_text.replace('ё', 'е')
def stem(text):
    # Разбиваем текст на отдельные слова
    words = text.split()
    # Стемминг каждого слова
    stemmed_words = [stemmer.stem(word) for word in words]
    # Склеиваем стеммированные слова обратно в текст
    stemmed_text = ' '.join(stemmed_words)
    return stemmed_text.replace('ё', 'е')
def remove_stop_words(text):
    words = text.split()  # Разбиваем текст на слова
    word_to_keep = "не"  # Слово, которое нужно сохранить
    filtered_words = [word for word in words
                      if word.lower() not in stop_words or word.lower() == word_to_keep.lower()]
    # Фильтруем слова, оставляя только те, которые не являются стоп-словами, за исключением слова "не"
    return ' '.join(filtered_words)  # Объединяем слова обратно в текст
#общие функции для обработки
def process_data_lem(data):
    data = remove_digits_and_punctuation(data)
    data = remove_stop_words(data)
    data = lemmatize(data)
    print("внутри лем ", data)
    return data
def process_data_stem(data):
    data = remove_digits_and_punctuation(data)
    data = remove_stop_words(data)
    data = stem(data)
    return data

# Функция для обработки текста с использованием токенизатора и модели LSTM
def process_text(input_text):
    if selected_model_path.endswith("LEM_model.keras"):
        input_text = process_data_lem(input_text)
        print("сейчас лем")
    else:
        input_text = process_data_stem(input_text)
        print("сейчас стем")


    # input_text=process_data_lem(input_text)
    # input_text=process_data_stem(input_text)

    # max_sequence_length = 704#1103
    # max_sequence_length = tokenizer.get_config()["max_sequence_length"]
    max_sequence_length = tokenizer.max_sequence_length
    # Токенизация текста
    input_sequence = tokenizer.texts_to_sequences([input_text])
    # Подготовка последовательности для модели (паддинг)
    padded_sequence = pad_sequences(input_sequence, maxlen=max_sequence_length)
    # Предсказание с помощью модели
    press = model.predict(padded_sequence)
    # Обработка результатов
    results = []
    for i in range(len(press)):
        sentiment_press1, sentiment_press2 = press[i][:2]
        sentiment_press = "Позитивный" if sentiment_press1 > sentiment_press2 else "Негативный"

        sentiment_price1, sentiment_price2 = press[i][2:4]
        sentiment_price = "Позитивный" if sentiment_price1 > sentiment_price2 else "Негативный"

        sentiment_sleep1, sentiment_sleep2 = press[i][4:6]
        sentiment_sleep = "Позитивный" if sentiment_sleep1 > sentiment_sleep2 else "Негативный"

        sentiment_numberH1, sentiment_numberH2 = press[i][6:8]
        sentiment_numberH = "Позитивный" if sentiment_numberH1 > sentiment_numberH2 else "Негативный"

        sentiment_clean1, sentiment_clean2 = press[i][8:10]
        sentiment_clean = "Позитивный" if sentiment_clean1 > sentiment_clean2 else "Негативный"

        sentiment_service1, sentiment_service2 = press[i][10:12]
        sentiment_service = "Позитивный" if sentiment_service1 > sentiment_service2 else "Негативный"


        results.append({
            # 'все': press[i],
            'вероятности общая': press[i][:2].tolist(),
            'Сентимент общий': sentiment_press,

            'вероятности цена/качество': press[i][2:4].tolist(),
            'Сентимент цена/качество': sentiment_price,

            'вероятности сон':  press[i][4:6].tolist(),
            'Сентимент сон': sentiment_sleep,

            'вероятности номер': press[i][6:8].tolist(),
            'Сентимент номер': sentiment_numberH,

            'вероятности чистота': press[i][8:10].tolist(),
            'Сентимент чистота': sentiment_clean,

            'вероятности сервис': press[i][10:12].tolist(),
            'Сентимент сервис': sentiment_service,

        })

    return press


def choose_doc():
    global selected_doc
    selected_doc = filedialog.askopenfilename()

    if selected_doc:
        # Проверка наличия файла и его расширения
        _, file_extension = os.path.splitext(selected_doc)
        if file_extension.lower() in ['.xlsx']:
            files_entry_doc.delete(0, tk.END)
            files_entry_doc.insert(0, selected_doc)
        else:
            messagebox.showwarning("Ошибка", "Неверное расширение файла. Выберите файл с расширением .xlsx")
            selected_doc = ""

def choose_model_path():
    global selected_model_path
    selected_model_path = filedialog.askopenfilename()

    if selected_model_path:
        # Проверка наличия файла и его расширения
        _, file_extension = os.path.splitext(selected_model_path)
        if file_extension.lower() in ['.keras']:
            files_entry_model.delete(0, tk.END)
            files_entry_model.insert(0, selected_model_path)
        else:
            messagebox.showwarning("Ошибка", "Неверное расширение файла. Выберите файл с расширением .keras")
            selected_model_path = ""
def choose_toki_path():
    global selected_toki_path
    selected_toki_path = filedialog.askopenfilename()

    if selected_toki_path:
        # Проверка наличия файла и его расширения
        _, file_extension = os.path.splitext(selected_toki_path)
        if file_extension.lower() in ['.pickle']:
            files_entry_toki.delete(0, tk.END)
            files_entry_toki.insert(0, selected_toki_path)
        else:
            messagebox.showwarning("Ошибка", "Неверное расширение файла. Выберите файл с расширением .pickle")
            selected_toki_path = ""
def load_model_and_tokenizer_if_paths_exist():
    global selected_model_path, selected_toki_path, model, tokenizer

    if selected_model_path is not None and selected_toki_path is not None:
        # Загрузка модели
        model = load_model(selected_model_path, custom_objects={'f1_m': f1_m, 'precision_m': precision_m, 'recall_m': recall_m})
        # Загрузка токенизатора
        with open(selected_toki_path, 'rb') as handle:
            tokenizer = pickle.load(handle)

#
def next_page_window():
    load_model_and_tokenizer_if_paths_exist()
    root.withdraw()  # Скрываем главное окно
    # Создание нового окна для следующей страницы интерфейса
    next_page_root = tk.Tk()
    next_page_root.title("Сентимент-анализ")
    next_page_root.geometry("1000x600")
    # Разделение окна на две части
    left_frame = tk.Frame(next_page_root, width=500, height=600)
    left_frame.pack_propagate(False)
    left_frame.pack(side=tk.LEFT)
    right_frame = tk.Frame(next_page_root, width=500, height=600)
    right_frame.pack_propagate(False)
    right_frame.pack(side=tk.RIGHT)

    # Поле ввода отзыва
    review_label = tk.Label(left_frame, text="Введите ваш отзыв:")
    review_label.pack(pady=10)
    # Создание виджета Text для ввода отзыва
    review_entry = tk.Text(left_frame, width=50, height=20, wrap=tk.WORD, font=("Times New Roman", 14), bd=2,
                           highlightthickness=1, highlightbackground="black")
    review_entry.bind("<Control-c>", copy_text)
    review_entry.bind("<Control-v>", paste_text)
    review_entry.pack()
    # Окно для отображения передаваемых значений
    values_label = tk.Label(right_frame, text="Сентимент-анализ:")
    values_label.pack(pady=10)
    values_display = tk.Text(right_frame, width=50, height=20, wrap=tk.WORD, font=("Times New Roman", 14), bd=2,
                           highlightthickness=1, highlightbackground="black")
    values_display.pack()

    def pass_value():
        values_display.delete("1.0", tk.END)
        # Обработка текста и вывод результатов
        value = review_entry.get("1.0", tk.END)  # Получаем текст из поля ввода
        processed_value = process_text(value)
        # Очищаем текст в поле для отображения
        values_display.delete("1.0", tk.END)
        values_display.tag_configure("green", foreground="green")
        values_display.tag_configure("red", foreground="red")
        # Выводим результаты из массива press
        for i in range(len(processed_value)):
            sentiment_press1, sentiment_press2 = processed_value[i][:2]
            sentiment_press = "Позитивный" if sentiment_press1 > sentiment_press2 else "Негативный"

            sentiment_price1, sentiment_price2 = processed_value[i][2:4]
            sentiment_price = "Позитивный" if sentiment_price1 > sentiment_price2 else "Негативный"

            sentiment_sleep1, sentiment_sleep2 = processed_value[i][4:6]
            sentiment_sleep = "Позитивный" if sentiment_sleep1 > sentiment_sleep2 else "Негативный"

            sentiment_numberH1, sentiment_numberH2 = processed_value[i][6:8]
            sentiment_numberH = "Позитивный" if sentiment_numberH1 > sentiment_numberH2 else "Негативный"

            sentiment_clean1, sentiment_clean2 = processed_value[i][8:10]
            sentiment_clean = "Позитивный" if sentiment_clean1 > sentiment_clean2 else "Негативный"

            sentiment_service1, sentiment_service2 = processed_value[i][10:12]
            sentiment_service = "Позитивный" if sentiment_service1 > sentiment_service2 else "Негативный"

            # Удаляем последние 20 цифр из каждого числа
            probabilities_general = [str(prob)[:25] for prob in processed_value[i][:2]]
            probabilities_price = [str(prob)[:25] for prob in processed_value[i][2:4]]
            probabilities_sleep = [str(prob)[:25] for prob in processed_value[i][4:6]]
            probabilities_number = [str(prob)[:25] for prob in processed_value[i][6:8]]
            probabilities_clean = [str(prob)[:25] for prob in processed_value[i][8:10]]
            probabilities_service = [str(prob)[:25] for prob in processed_value[i][10:12]]

            color_press = "green" if sentiment_press == "Позитивный" else "red"
            color_price = "green" if sentiment_price== "Позитивный" else "red"
            color_sleep = "green" if sentiment_sleep == "Позитивный" else "red"
            color_numberH = "green" if sentiment_numberH == "Позитивный" else "red"
            color_clean = "green" if sentiment_clean == "Позитивный" else "red"
            color_service = "green" if sentiment_service == "Позитивный" else "red"

            values_display.insert(tk.END, f"вероятности общая: ")
            values_display.insert(tk.END, f"{probabilities_general[0]} ", "green")
            values_display.insert(tk.END, f"{probabilities_general[1]} ", "red")  # Вставляем число с применением соответствующего тега
            # Вставляем число с применением соответствующего тега
            values_display.insert(tk.END, f"\n")
            values_display.insert(tk.END, f"Сентимент общий: ")
            values_display.insert(tk.END, f"{sentiment_press}\n", color_press)

            values_display.insert(tk.END, f"вероятности цена/качество: ")
            values_display.insert(tk.END, f"{probabilities_price[0]} ", "green")
            values_display.insert(tk.END, f"{probabilities_price[1]} ", "red")
            values_display.insert(tk.END, f"\n")
            values_display.insert(tk.END, f"Сентимент цена/качество: ")
            values_display.insert(tk.END, f"{sentiment_price}\n", color_price)

            values_display.insert(tk.END, f"вероятности сон: ")
            values_display.insert(tk.END, f"{probabilities_sleep[0]} ", "green")
            values_display.insert(tk.END, f"{probabilities_sleep[1]} ", "red")
            values_display.insert(tk.END, f"\n")
            values_display.insert(tk.END, f"Сентимент сон: ")
            values_display.insert(tk.END, f"{sentiment_sleep}\n", color_sleep)

            values_display.insert(tk.END, f"вероятности номер: ")
            values_display.insert(tk.END, f"{probabilities_number[0]} ", "green")
            values_display.insert(tk.END, f"{probabilities_number[1]} ", "red")
            values_display.insert(tk.END, f"\n")
            values_display.insert(tk.END, f"Сентимент номер: ")
            values_display.insert(tk.END, f"{sentiment_numberH}\n", color_numberH)

            values_display.insert(tk.END, f"вероятности чистота: ")
            values_display.insert(tk.END, f"{probabilities_clean[0]} ", "green")
            values_display.insert(tk.END, f"{probabilities_clean[1]} ", "red")
            values_display.insert(tk.END, f"\n")
            values_display.insert(tk.END, f"Сентимент чистота: ")
            values_display.insert(tk.END, f"{sentiment_clean}\n", color_clean)

            values_display.insert(tk.END, f"вероятности обслужвание: ")
            values_display.insert(tk.END, f"{probabilities_service[0]} ", "green")
            values_display.insert(tk.END, f"{probabilities_service[1]} ", "red")
            values_display.insert(tk.END, f"\n")
            values_display.insert(tk.END, f"Сентимент обслужвание: ")
            values_display.insert(tk.END, f"{sentiment_service}\n", color_service)

            values_display.insert(tk.END, "\n")

    # Кнопка для передачи значения
    pass_button = tk.Button(left_frame, text="Выполнить", command=pass_value)
    pass_button.pack(pady=10)

    # Кнопка для возвращения к первому окну
    back_button = tk.Button(left_frame, text="Назад", command=lambda: back_to_main(next_page_root))
    back_button.pack(side=tk.BOTTOM, pady=10)

    next_page_root.mainloop()

def next_page_window_doc():
    messagebox.showwarning("Загрузка", "Сентимент-анализ выполняется")
    load_model_and_tokenizer_if_paths_exist()
    df = pd.read_excel(selected_doc)
    text_array = df.iloc[:, 0].astype(str).tolist()
    max_sequence_length = 704

    jojo = [remove_digits_and_punctuation(text) for text in text_array]
    jojo = [remove_stop_words(text) for text in jojo]
    jojo = [lemmatize(text) for text in jojo]
    data123 = tokenizer.texts_to_sequences(jojo)
    data_pad1 = pad_sequences(data123, maxlen=max_sequence_length)
    press = model.predict(data_pad1)
    press_sent = press[:, :2]  # сентимент
    press_price = press[:, 2:4]
    press_sleep = press[:, 4:6]
    press_numberH = press[:, 6:8]
    press_clean = press[:, 8:10]
    press_service = press[:, 10:12]
    print(len(jojo))
    rows = []
    for i in range(len(jojo)):
        sentiment_press = "Позитивный" if press_sent[i][0] > press_sent[i][1] else "Негативный"
        sentiment_price = "Позитивный" if press_price[i][0] > press_price[i][1] else "Негативный"
        sentiment_sleep = "Позитивный" if press_sleep[i][0] > press_sleep[i][1] else "Негативный"
        sentiment_numberH = "Позитивный" if press_numberH[i][0] > press_numberH[i][1] else "Негативный"
        sentiment_clean = "Позитивный" if press_clean[i][0] > press_clean[i][1] else "Негативный"
        sentiment_service = "Позитивный" if press_service[i][0] > press_service[i][1] else "Негативный"

        rows.append({
            'вероятность общая': press_sent[i],
            'Сентимент общий': sentiment_press,

            'вероятности цена/качество': press_price[i],
            'Сентимент цена/качество': sentiment_price,

            'вероятности сон': press_sleep[i],
            'Сентимент сон': sentiment_sleep,

            'вероятности номер': press_numberH[i],
            'Сентимент номер': sentiment_numberH,

            'вероятности чистота': press_clean[i],
            'Сентимент чистота': sentiment_clean,

            'вероятности обслуживание': press_service[i],
            'Сентимент обслуживание': sentiment_service,

            'текст': text_array[i]
        })
    df = pd.DataFrame(rows)
    source_file_path = files_entry_doc.get()
    directory, filename = os.path.split(source_file_path)
    new_excel_file_path = os.path.join(directory, "Sentiment_new_" + filename)
    # excel_file_path = 'C:\\Users\\user\\PycharmProjects\\sentimenLSTM\\LSTM\\LSTM_LEM\\final\\text_full.xlsx'
    df.to_excel(new_excel_file_path, index=False)

    def adjust_column_width(sheet):
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    def color_cells(sheet, df):
        positive_font = Font(color="00FF00")
        negative_font = Font(color="FF0000")

        for row in sheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.column in [df.columns.get_loc('Сентимент общий') + 1,
                                   df.columns.get_loc('Сентимент цена/качество') + 1,
                                   df.columns.get_loc('Сентимент сон') + 1,
                                   df.columns.get_loc('Сентимент номер') + 1,
                                   df.columns.get_loc('Сентимент чистота') + 1,
                                   df.columns.get_loc('Сентимент обслуживание') + 1]:
                    if cell.value == "Позитивный":
                        cell.font = positive_font
                    elif cell.value == "Негативный":
                        cell.font = negative_font

    wb = load_workbook(new_excel_file_path)
    ws = wb.active
    adjust_column_width(ws)
    color_cells(ws, df)
    wb.save(new_excel_file_path)

    messagebox.showwarning("Файл", "Файл сохранен")

def choose_model_path_for_model(entry,index):
    global selected_model_paths_for_model
    new_model_path = filedialog.askopenfilename()

    if new_model_path:
        _, file_extension = os.path.splitext(new_model_path)
        if file_extension.lower() in ['.keras']:
            if len(selected_model_paths_for_model) > index:
                selected_model_paths_for_model[index] = new_model_path
            else:
                selected_model_paths_for_model.append(new_model_path)
            entry.delete(0, tk.END)
            entry.insert(0, new_model_path)
            # selected_model_paths_for_model.append(new_model_path)
            # entry.delete(0, tk.END)
            # entry.insert(0, '\n'.join(selected_model_paths_for_model))

        else:
            messagebox.showwarning("Ошибка", "Неверное расширение файла. Выберите файл с расширением .keras")

def choose_toki_path_for_model(entry,index):
    global selected_toki_path_for_model
    selected_toki_path = filedialog.askopenfilename()

    if selected_toki_path:
        # Проверка наличия файла и его расширения
        _, file_extension = os.path.splitext(selected_toki_path)
        if file_extension.lower() in ['.pickle']:
            if len(selected_toki_path_for_model) > index:
                selected_toki_path_for_model[index] = selected_toki_path
            else:
                selected_toki_path_for_model.append(selected_toki_path)
            entry.delete(0, tk.END)
            entry.insert(0, selected_toki_path)
            # selected_toki_path_for_model.append(selected_toki_path)
            # entry.delete(0, tk.END)
            # entry.insert(0, '\n'.join(selected_toki_path_for_model))
        else:
            messagebox.showwarning("Ошибка", "Неверное расширение файла. Выберите файл с расширением .pickle")

def load_models_and_tokenizers_if_paths_exist_for_model():
    global selected_model_paths_for_model, selected_toki_path_for_model, models, tokenizers

    if selected_model_paths_for_model and selected_toki_path_for_model:
        # Загрузка токенизаторов
        tokenizers = []
        for toki_path in selected_toki_path_for_model:
            with open(toki_path, 'rb') as handle:
                tokenizer = pickle.load(handle)
                tokenizers.append(tokenizer)

        # Загрузка моделей
        models = []
        for model_path in selected_model_paths_for_model:
            model = load_model(model_path, custom_objects={'f1_m': f1_m, 'precision_m': precision_m, 'recall_m': recall_m})
            models.append(model)

        print("Пг:", models)

        print("Пути к загруженным моделям:")
        for model_path in selected_model_paths_for_model:
            print(model_path)

        print("\nПути к загруженным токенизаторам:")
        for toki_path in selected_toki_path_for_model:
            print(toki_path)

def process_text_for_models(input_text):

    input_text = process_data_lem(input_text)

    # max_sequence_length = 704#1103
    # max_sequence_length = tokenizer.get_config()["max_sequence_length"]
    # max_sequence_length = tokenizer.max_sequence_length
    # # Токенизация текста
    # input_sequence = tokenizer.texts_to_sequences([input_text])
    # # Подготовка последовательности для модели (паддинг)
    # padded_sequence = pad_sequences(input_sequence, maxlen=max_sequence_length)
    # # Предсказание с помощью модели
    # press = model.predict(padded_sequence)
    # # Обработка результатов
    results = []
    for model, tokenizer in zip(models, tokenizers):
        max_sequence_length = tokenizer.max_sequence_length

        # Токенизация текста
        input_sequence = tokenizer.texts_to_sequences([input_text])

        # Подготовка последовательности для модели (паддинг)
        padded_sequence = pad_sequences(input_sequence, maxlen=max_sequence_length)

        # Предсказание с помощью модели
        press = model.predict(padded_sequence)

        # Обработка результатов и добавление в список результатов
        results.append(press)
    # for i, result in enumerate(results):
    #     print(f"Результаты для модели {i + 1}:")
    #     print(result)
    return results


def next_page_window_for_model_next():
    global selected_model_paths_for_model
    load_models_and_tokenizers_if_paths_exist_for_model()
    root.withdraw()  # Скрываем главное окно

    # Создание нового окна для следующей страницы интерфейса
    next_page_root = tk.Toplevel()
    next_page_root.title("Сентимент-анализ для сравнения моделей")
    next_page_root.geometry("1600x600")
    first_frame = tk.Frame(next_page_root, width=500, height=600)
    first_frame.pack_propagate(False)
    first_frame.pack(side=tk.LEFT)
    # Создание четырех фреймов для каждой модели
    frames = []
    for i in range(4):
        frame = tk.Frame(next_page_root, width=250, height=600)
        frame.pack_propagate(False)
        frame.pack(side=tk.LEFT)
        frames.append(frame)

    # Поле ввода отзыва
    review_label = tk.Label(first_frame, text="Введите ваш отзыв:")
    review_label.pack(pady=10)
    # Создание виджета Text для ввода отзыва
    review_entry = tk.Text(first_frame, width=100, height=20, wrap=tk.WORD,
                           font=("Times New Roman", 14), bd=2,
                           highlightthickness=1, highlightbackground="black")
    review_entry.bind("<Control-c>", copy_text)
    review_entry.bind("<Control-v>", paste_text)
    review_entry.pack()
    # Создание окон для отображения передаваемых значений для каждой модели
    values_labels = []
    values_displays = []
    for i in range(4):
        values_label = tk.Label(frames[i], text=f"Сентимент-анализ модели {i + 1}:")
        values_label.pack(pady=10)
        values_labels.append(values_label)
        values_display = tk.Text(frames[i], width=50, height=20, wrap=tk.WORD, font=("Times New Roman", 14), bd=2,
                                 highlightthickness=1, highlightbackground="black")
        values_display.pack()
        values_displays.append(values_display)

    def pass_value():
        
        selected_model_names = [model_path.split('/')[-1].split('.')[0] for model_path in selected_model_paths_for_model]

        value = review_entry.get("1.0", tk.END)
        processed_value = process_text_for_models(value)
        for j, result in enumerate(processed_value):
            print(f"Результаты для модели {j + 1}:")
            print(result)
        for i in range(4):
            values_displays[i].delete("1.0", tk.END)

            values_displays[i].tag_configure("green", foreground="green")
            values_displays[i].tag_configure("red", foreground="red")

            model_results = processed_value[i]  # Результаты только для текущей модели
            # Очищаем текст в поле для отображения
            values_displays[i].delete("1.0", tk.END)
            values_displays[i].tag_configure("green", foreground="green")
            values_displays[i].tag_configure("red", foreground="red")
            words = ["общий", "цена/качество", "сон", "номер", "чистота", "обслуживание"]
            model_name = selected_model_names[i]
            values_displays[i].insert(tk.END, f"Результаты для модели {model_name}:\n")

            for k in range(0, len(model_results[0]), 2):
                probabilities = model_results[0][k:k + 2]
                sentiment = ["Позитивный" if prob1 > prob2 else "Негативный" for prob1, prob2 in
                             zip(probabilities[::2], probabilities[1::2])]
                colors = ["green" if sent == "Позитивный" else "red" for sent in sentiment]
                for l, (prob1, prob2, sent, color) in enumerate(
                        zip(probabilities[::2], probabilities[1::2], sentiment, colors)):
                    word = words[k//2]
                    values_displays[i].insert(tk.END, f"Вероятности {word}:\n ")
                    values_displays[i].insert(tk.END, f"{prob1:.7f}, ", "green")
                    values_displays[i].insert(tk.END, f"{prob2:.7f}\n", "red")
                    values_displays[i].insert(tk.END, "Сентимент: ")
                    values_displays[i].insert(tk.END, f"{sent}\n", color)
            values_displays[i].insert(tk.END, "\n")

    pass_button = tk.Button(first_frame, text="Выполнить", command=pass_value)
    pass_button.pack(pady=10)

    # Кнопка для возвращения к первому окну
    back_button = tk.Button(frames[1], text="Назад", command=lambda: back_to_main(next_page_root))
    back_button.pack(pady=20,anchor='center')

    next_page_root.mainloop()


def next_page_window_for_model():
    global files_entry_for_models
    root.withdraw()  # Скрываем главное окно

    next_page_window_for_model = tk.Tk()
    next_page_window_for_model.title("Сентимент-анализ с сравнением моделей")
    next_page_window_for_model.geometry("800x600")
    frame_models = tk.Frame(next_page_window_for_model)
    frame_models.pack(padx=20, pady=20)


#Окно для пользовательского ввода путей к файлам моделей keras и токинезаторов pickle
#
#1 модель
#
    files_label1_0 = tk.Label(frame_models, text="Путь к 1-ой модели:")
    files_label1_0.grid(row=1, column=0, sticky="w")
    files_entry_for_models1_0 = tk.Entry(frame_models, width=70)
    files_entry_for_models1_0.grid(row=1, column=1, padx=10, pady=5)
    if not selected_model_paths_for_model:
        files_button = tk.Button(frame_models, text="Выбрать", command=lambda: choose_model_path_for_model(files_entry_for_models1_0,0))
        files_button.grid(row=1, column=2, padx=10, pady=5)
    else:
        files_entry_for_models1_0.insert(tk.END, selected_model_paths_for_model[0])
        print("Путь к модели 1-ой модели:",selected_model_paths_for_model[0])
        files_button = tk.Button(frame_models, text="Выбрать", command=lambda: choose_model_path_for_model(files_entry_for_models1_0,0))
        files_button.grid(row=1, column=2, padx=10, pady=5)
#1 модель токинезатор
    files_label1_1 = tk.Label(frame_models, text="Путь к токенизатору 1-ой модели:")
    files_label1_1.grid(row=2, column=0, sticky="w")
    files_entry_for_toki1_1 = tk.Entry(frame_models, width=70)
    files_entry_for_toki1_1.grid(row=2, column=1, padx=10, pady=5)
    if not selected_toki_path_for_model:
        files_button1 = tk.Button(frame_models, text="Выбрать", command=lambda: choose_toki_path_for_model(files_entry_for_toki1_1,0))
        files_button1.grid(row=2, column=2, padx=10, pady=5)
    else:
        files_entry_for_toki1_1.insert(tk.END, selected_toki_path_for_model[0])
        print("Путь к токенизатору 1-ой модели:",selected_toki_path_for_model[0])
        files_button1 = tk.Button(frame_models, text="Выбрать", command=lambda: choose_toki_path_for_model(files_entry_for_toki1_1,0))
        files_button1.grid(row=2, column=2, padx=10, pady=5)
#
#2 модель
#
    files_label2_1 = tk.Label(frame_models, text="Путь к 2-ой модели:")
    files_label2_1.grid(row=3, column=0, sticky="w")
    files_entry_for_models2_0 = tk.Entry(frame_models, width=70)
    files_entry_for_models2_0.grid(row=3, column=1, padx=10, pady=5)
    if not selected_model_paths_for_model:
        files_button = tk.Button(frame_models, text="Выбрать", command=lambda: choose_model_path_for_model(files_entry_for_models2_0,1))
        files_button.grid(row=3, column=2, padx=10, pady=5)
    else:
        files_entry_for_models2_0.insert(tk.END, selected_model_paths_for_model[1])
        print("Путь к модели 2-ой модели:", selected_model_paths_for_model[1])
        files_button = tk.Button(frame_models, text="Выбрать",
                                 command=lambda: choose_model_path_for_model(files_entry_for_models2_0, 1))
        files_button.grid(row=3, column=2, padx=10, pady=5)
#2 модель токинезатор
    files_label1 = tk.Label(frame_models, text="Путь к токенизатору 2-ой модели:")
    files_label1.grid(row=4, column=0, sticky="w")
    files_entry_toki2_1 = tk.Entry(frame_models, width=70)
    files_entry_toki2_1.grid(row=4, column=1, padx=10, pady=5)
    if not selected_toki_path_for_model:
        files_button1 = tk.Button(frame_models, text="Выбрать", command=lambda: choose_toki_path_for_model(files_entry_toki2_1,1))
        files_button1.grid(row=4, column=2, padx=10, pady=5)
    else:
        files_entry_toki2_1.insert(tk.END, selected_toki_path_for_model[1])
        print("Путь к токенизатору 2-ой модели:", selected_toki_path_for_model[1])
        files_button1 = tk.Button(frame_models, text="Выбрать",
                                  command=lambda: choose_toki_path_for_model(files_entry_toki2_1, 1))
        files_button1.grid(row=4, column=2, padx=10, pady=5)
#
#3 модель
#
    files_label3_0 = tk.Label(frame_models, text="Путь к 3-ей модели:")
    files_label3_0.grid(row=5, column=0, sticky="w")
    files_entry_model3_0 = tk.Entry(frame_models, width=70)
    files_entry_model3_0.grid(row=5, column=1, padx=10, pady=5)
    if not selected_model_paths_for_model:
        files_button = tk.Button(frame_models, text="Выбрать", command=lambda: choose_model_path_for_model(files_entry_model3_0,2))
        files_button.grid(row=5, column=2, padx=10, pady=5)
    else:
        files_entry_model3_0.insert(tk.END, selected_model_paths_for_model[2])
        print("Путь к модели 3-ой модели:",selected_model_paths_for_model[2])
        files_button = tk.Button(frame_models, text="Выбрать",
                                 command=lambda: choose_model_path_for_model(files_entry_model3_0, 2))
        files_button.grid(row=5, column=2, padx=10, pady=5)
#3 модель токинезатор
    files_label3_1 = tk.Label(frame_models, text="Путь к токенизатору 3-ей модели:")
    files_label3_1.grid(row=6, column=0, sticky="w")
    files_entry_toki3_1 = tk.Entry(frame_models, width=70)
    files_entry_toki3_1.grid(row=6, column=1, padx=10, pady=5)
    if not selected_toki_path_for_model:
        files_button1 = tk.Button(frame_models, text="Выбрать", command=lambda: choose_toki_path_for_model(files_entry_toki3_1,2))
        files_button1.grid(row=6, column=2, padx=10, pady=5)
    else:
        files_entry_toki3_1.insert(tk.END, selected_toki_path_for_model[2])
        print("Путь к токенизатору 1-ой модели:",selected_toki_path_for_model[2])
        files_button1 = tk.Button(frame_models, text="Выбрать",
                                  command=lambda: choose_toki_path_for_model(files_entry_toki3_1, 2))
        files_button1.grid(row=6, column=2, padx=10, pady=5)
#
#4 модель
#
    files_label4_1 = tk.Label(frame_models, text="Путь к 4-ой модели:")
    files_label4_1.grid(row=7, column=0, sticky="w")
    files_entry_model4_0 = tk.Entry(frame_models, width=70)
    files_entry_model4_0.grid(row=7, column=1, padx=10, pady=5)
    if not selected_model_paths_for_model:
        files_button = tk.Button(frame_models, text="Выбрать", command=lambda: choose_model_path_for_model(files_entry_model4_0,3))
        files_button.grid(row=7, column=2, padx=10, pady=5)
    else:
        files_entry_model4_0.insert(tk.END, selected_model_paths_for_model[3])
        print("Путь к модели 4-ой модели:", selected_model_paths_for_model[3])
        files_button = tk.Button(frame_models, text="Выбрать",
                                 command=lambda: choose_model_path_for_model(files_entry_model4_0, 3))
        files_button.grid(row=7, column=2, padx=10, pady=5)
#4 модель токинезатор
    files_label4_0 = tk.Label(frame_models, text="Путь к токенизатору 4-ой модели:")
    files_label4_0.grid(row=8, column=0, sticky="w")
    files_entry_toki4_1 = tk.Entry(frame_models, width=70)
    files_entry_toki4_1.grid(row=8, column=1, padx=10, pady=5)
    if not selected_toki_path_for_model:
        files_button1 = tk.Button(frame_models, text="Выбрать", command=lambda: choose_toki_path_for_model(files_entry_toki4_1,3))
        files_button1.grid(row=8, column=2, padx=10, pady=5)
    else:
        files_entry_toki4_1.insert(tk.END, selected_toki_path_for_model[3])
        print("Путь к токенизатору 4-ой модели:", selected_toki_path_for_model[3])
        files_button1 = tk.Button(frame_models, text="Выбрать",
                                  command=lambda: choose_toki_path_for_model(files_entry_toki4_1, 3))
        files_button1.grid(row=8, column=2, padx=10, pady=5)

    # Кнопка для перехода на второе окно
    next_window_button = tk.Button(frame_models, text="Загрузка моделей", command=next_page_window_for_model_next)
    next_window_button.grid(row=9, column=1, pady=10)
    # Кнопка для возвращения к первому окну
    # Кнопка для возвращения к первому окну
    back_button = tk.Button(frame_models, text="Назад", command=lambda: back_to_main(next_page_window_for_model))
    back_button.grid(row=10, column=1, pady=10)

    next_page_window_for_model.mainloop()

def back_to_main(window):
    window.destroy()
    root.deiconify()
def exit_program():
    root.destroy()
def copy_text(event):
    if event.state & 4 and event.keysym == 'c':
        # if event.state & 4 and event.keysym == 'c' or event.keysym == 'с':
        event.widget.event_generate("<<Copy>>")
        return "break"  # Предотвращает дальнейшую обработку события
    elif event.state & 4 and event.keysym == 'v':
        event.widget.event_generate("<<Paste>>")
        return "break"  # Предотвращает дальнейшую обработку события
def paste_text(event):
    if event.state & 4 and event.keysym == 'v':
        event.widget.event_generate("<<Paste>>")
        return "break"  # Предотвращает дальнейшую обработку события

def main():
    global root, files_entry_model, files_entry_toki, files_entry_doc


    root = tk.Tk()
    root.title("Сентимент-анализ")
    root.geometry("600x350")

    # Создание рамки
    frame = tk.Frame(root)
    frame.pack(padx=20, pady=20)

    # Пользовательский ввод для пути к файлам
    files_label = tk.Label(frame, text="Путь к модели:")
    files_label.grid(row=1, column=0, sticky="w")

    files_entry_model = tk.Entry(frame, width=50)
    files_entry_model.grid(row=1, column=1, padx=10, pady=5)

    if selected_model_path is None:
        files_button = tk.Button(frame, text="Выбрать", command=choose_model_path)
        files_button.grid(row=1, column=2, padx=10, pady=5)
    else:
        files_entry_model.insert(tk.END, selected_model_path)

    # Пользовательский ввод для пути к файлам
    files_label = tk.Label(frame, text="Путь к токенизатору:")
    files_label.grid(row=2, column=0, sticky="w")

    files_entry_toki = tk.Entry(frame, width=50)
    files_entry_toki.grid(row=2, column=1, padx=10, pady=5)

    if selected_toki_path is None:
        files_button = tk.Button(frame, text="Выбрать", command=choose_toki_path)
        files_button.grid(row=2, column=2, padx=10, pady=5)
    else:
        files_entry_toki.insert(tk.END, selected_toki_path)

    # Кнопка для перехода на второе окно
    next_window_button = tk.Button(frame, text="Загрузка модели", command=next_page_window)
    next_window_button.grid(row=3, column=1, pady=10)

    # Пользовательский ввод документа
    files_label = tk.Label(frame, text="Путь к документу:")
    files_label.grid(row=4, column=0, sticky="w")

    files_entry_doc = tk.Entry(frame, width=50)
    files_entry_doc.grid(row=4, column=1, padx=10, pady=5)

    if selected_doc is None:
        files_button = tk.Button(frame, text="Выбрать", command=choose_doc)
        files_button.grid(row=4, column=2, padx=10, pady=5)
    else:
        files_entry_doc.insert(tk.END, selected_doc)
    # # Кнопка для загрузки документа
    load_button = tk.Button(frame, text="Загрузить документ", command=next_page_window_doc)
    load_button.grid(row=5, column=1, padx=10, pady=5)

    load_button = tk.Button(frame, text="Сравнение моделей", command=next_page_window_for_model)
    load_button.grid(row=6, column=1, padx=10, pady=15)





    # Кнопка для выхода из программы
    exit_button = tk.Button(root, text="Выход", command=exit_program)
    exit_button.pack(side=tk.BOTTOM, pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
    if selected_model_path:
        # Здесь вы можете использовать путь к выбранной модели для загрузки модели и дальнейшей работы с ней во втором окне
        print("Выбранный файл модели:", selected_model_path)
    if selected_toki_path:
        # Здесь вы можете использовать путь к выбранной модели для загрузки модели и дальнейшей работы с ней во втором окне
        print("Выбранный файл модели:", selected_toki_path)
    # for result in results:
    #     for key, value in result.items():
    #         print(f'{key}: {value}')
    #     print()